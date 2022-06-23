import subprocess
import time
import win32gui, win32con, win32api #pywin32
import pyautogui
from openpyxl import load_workbook, Workbook
import openpyxl
import pyperclip
from PIL import ImageGrab #pillow
import re
import os
from openpyxl.styles import PatternFill

from PyQt5.QtWidgets import (QApplication, 
        QDialog, QGridLayout, QHBoxLayout, QLabel, QLineEdit,
        QPushButton, QStyleFactory, QFileDialog)
from PyQt5.QtGui import QPixmap, QFont
import sys


class SimFCSAutomator(QDialog):
    def __init__(self, parent=None):
        super(SimFCSAutomator, self).__init__(parent)

        self.originalPalette = QApplication.palette()

        # Define colour to fill Excel cell with (for unique numbers)
        self.fill_green = PatternFill(patternType='solid', fgColor='35FC03')
        self.fill_yellow = PatternFill(patternType='solid', fgColor='FFFF00')
        self.fill_red = PatternFill(patternType='solid', fgColor='FF0000')

        # Input logo images
        logoCI = QLabel(self)
        pixmap = QPixmap("Centenary_Institute_logo.png")
        logoCI.setPixmap(pixmap.scaled(200, 120))
        logoGIC = QLabel(self)
        pixmap = QPixmap("GIC.png")
        logoGIC.setPixmap(pixmap.scaled(200, 120))

        # Input Dashboard Title
        dashboardTitle = QLabel("SimFCS Automator")
        dashboardTitle.setFont(QFont("Arial", 24))

        topLayout = QHBoxLayout()
        topLayout.addStretch()
        topLayout.addWidget(logoCI)
        topLayout.addWidget(dashboardTitle)
        topLayout.addWidget(logoGIC)
        topLayout.addStretch()

        self.fileLocButton = QPushButton("&Data Folder")
        self.fileLocText = QLabel()
        self.fileLocButton.pressed.connect(self.getFileLoc)

        self.imageLocButton = QPushButton("&Images Folder")
        self.imageLocText = QLabel()
        self.imageLocButton.pressed.connect(self.getImageLoc)

        self.excelLocButton = QPushButton("&Excel Folder")
        self.excelLocText = QLabel()
        self.excelLocButton.pressed.connect(self.getExcelLoc)

        self.excelFileNameText = QLabel()
        self.excelFileNameText.setText("Excel File Name:")
        self.excelFileName = QLineEdit(self)
        self.excelFileName.setText("Excel")

        self.simFCSButton = QPushButton("&SimFCS Location")
        self.simFCSText = QLabel()
        self.simFCSButton.pressed.connect(self.getSimFCS)
        # self.simFCSButton.pressed.connect(partial(self.getSimFCS, test))

        self.shortWaitText = QLabel()
        self.shortWaitText.setText("Short:")
        self.shortWaitBox = QLineEdit(self)
        self.shortWaitBox.setText("0.1")
        self.mediumWaitText = QLabel()
        self.mediumWaitText.setText("Medium:")
        self.mediumWaitBox = QLineEdit(self)
        self.mediumWaitBox.setText("0.5")
        self.longWaitText = QLabel()
        self.longWaitText.setText("Long:")
        self.longWaitBox = QLineEdit(self)
        self.longWaitBox.setText("2.0")
        self.simulationWaitText = QLabel()
        self.simulationWaitText.setText("Simulation:")
        self.simulationWaitBox = QLineEdit(self)
        self.simulationWaitBox.setText("10.0")

        self.runButton = QPushButton("&Run")
        self.runButton.pressed.connect(self.startSimFCS)

        mainLayout = QGridLayout()
        mainLayout.addLayout(topLayout, 0, 0, 1, 3) # x, y, x-span, y-span
        mainLayout.addWidget(self.fileLocButton, 1, 0)
        mainLayout.addWidget(self.fileLocText, 1, 1)
        mainLayout.addWidget(self.imageLocButton, 2, 0)
        mainLayout.addWidget(self.imageLocText, 2, 1)
        mainLayout.addWidget(self.excelLocButton, 3, 0)
        mainLayout.addWidget(self.excelLocText, 3, 1)
        mainLayout.addWidget(self.excelFileNameText, 4, 1)
        mainLayout.addWidget(self.excelFileName, 4, 2)
        mainLayout.addWidget(self.simFCSButton, 5, 0)
        mainLayout.addWidget(self.simFCSText, 5, 1)

        mainLayout.addWidget(self.shortWaitText, 7, 0)
        mainLayout.addWidget(self.shortWaitBox, 7, 1)
        mainLayout.addWidget(self.mediumWaitText, 8, 0)
        mainLayout.addWidget(self.mediumWaitBox, 8, 1)
        mainLayout.addWidget(self.longWaitText, 9, 0)
        mainLayout.addWidget(self.longWaitBox, 9, 1)
        mainLayout.addWidget(self.simulationWaitText, 10, 0)
        mainLayout.addWidget(self.simulationWaitBox, 10, 1)
        mainLayout.addWidget(self.runButton, 11, 1, 1, 3)
        mainLayout.setRowStretch(1, 6) # stretching which row, by how much
        mainLayout.setRowStretch(2, 1)
        mainLayout.setRowStretch(3, 1)
        mainLayout.setRowStretch(4, 1)
        mainLayout.setRowStretch(5, 1)
        mainLayout.setRowStretch(6, 1)
        mainLayout.setRowStretch(7, 1)
        mainLayout.setRowStretch(8, 1)
        mainLayout.setRowStretch(9, 1)
        mainLayout.setRowStretch(10, 1)
        mainLayout.setRowStretch(11, 1)
        mainLayout.setColumnStretch(0, 1)
        mainLayout.setColumnStretch(1, 2)
        mainLayout.setColumnStretch(2, 2)
        self.setLayout(mainLayout)

        self.setWindowTitle("SimFCS Automator")
        self.changeStyle('Fusion')

    def getFileLoc(self):
        self.fileLoc = QFileDialog.getExistingDirectory(self, "Choose Directory of Data Folders","C:/")
        self.fileLocText.setText(self.fileLoc)

    def getImageLoc(self):
        self.imageLoc = QFileDialog.getExistingDirectory(self, "Choose Directory for Image Output","C:/")
        self.imageLocText.setText(self.imageLoc)

    def getExcelLoc(self):
        self.excelFileDir = QFileDialog.getExistingDirectory(self, "Choose Directory for Excel Output","C:/")
        self.excelLocText.setText(self.excelFileDir)

    def getSimFCS(self):
        simFCSDir = QFileDialog.getOpenFileNames(self,"Choose File","C:/")
        self.simFCSLoc = simFCSDir[0][0]
        self.simFCSText.setText(self.simFCSLoc)

    def changeStyle(self, styleName):
        QApplication.setStyle(QStyleFactory.create(styleName))
        self.changePalette()

    def changePalette(self):
        QApplication.setPalette(QApplication.style().standardPalette())
        QApplication.setPalette(self.originalPalette)

    def click(self, x,y):
        win32api.SetCursorPos((x,y))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,x,y,0,0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,x,y,0,0)

    def selectFiles(self):
        pyautogui.keyDown('alt')
        pyautogui.press('f')
        time.sleep(self.mediumWait)
        pyautogui.press('p')
        pyautogui.keyUp('alt')
        time.sleep(self.mediumWait)

    def exchangeChannel(self):
        pyautogui.keyDown('alt')
        pyautogui.press('t')
        time.sleep(self.mediumWait)
        pyautogui.press('e')
        pyautogui.keyUp('alt')
        time.sleep(self.mediumWait)

    def selectGroup(self, fileDir, group):
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        pyautogui.press('enter')
        time.sleep(self.shortWait)
        pyautogui.typewrite(fileDir)
        time.sleep(self.mediumWait)
        pyautogui.press('enter')
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        pyautogui.press('tab')
        time.sleep(self.shortWait)
        time.sleep(self.shortWait)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.shortWait)
        pyautogui.press('backspace')
        time.sleep(self.shortWait)
        pyautogui.typewrite(group)
        time.sleep(self.shortWait)
        pyautogui.press('enter')
        time.sleep(self.shortWait)
        self.click(341,284)
        time.sleep(self.shortWait)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.shortWait)
        pyautogui.press('enter')

    def substractMovingAverage(self):
        pyautogui.keyDown('alt')
        pyautogui.press('t')
        time.sleep(self.mediumWait)
        pyautogui.press('r')
        time.sleep(self.mediumWait)
        pyautogui.press('s')
        pyautogui.keyUp('alt')
        time.sleep(self.longWait)

    def takeImages(self, filename, channel):
        if channel == 0:
            # Take first image
            pyautogui.moveTo(147, 200)
            pyautogui.click(button='right')
            time.sleep(self.mediumWait)
            self.click(219,212)
            time.sleep(self.mediumWait)
            pyautogui.keyDown('alt')
            time.sleep(self.shortWait)
            pyautogui.press('c')
            time.sleep(self.mediumWait)
            pyautogui.keyUp('alt')
            time.sleep(self.mediumWait)
            selectWin = win32gui.GetForegroundWindow()
            while win32gui.GetWindowText(selectWin) != "":
                time.sleep(self.longWait)
                selectWin = win32gui.GetForegroundWindow()
            win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

            img = ImageGrab.grabclipboard()
            img.save(self.imageLoc + filename + str(channel) + '1.png')
            img=openpyxl.drawing.image.Image(self.imageLoc + filename + str(channel) + '1.png')
            img.anchor = self.imageCol[channel * 2] + str(self.rowToImage)
            self.wsI.add_image(img)
            time.sleep(self.longWait)
        elif channel == 1:
            # Take first image
            pyautogui.moveTo(471, 205)
            pyautogui.click(button='right')
            time.sleep(self.mediumWait)
            self.click(511,218)
            time.sleep(self.mediumWait)
            pyautogui.keyDown('alt')
            time.sleep(self.shortWait)
            pyautogui.press('c')
            time.sleep(self.mediumWait)
            pyautogui.keyUp('alt')
            time.sleep(self.mediumWait)
            selectWin = win32gui.GetForegroundWindow()
            while win32gui.GetWindowText(selectWin) != "":
                time.sleep(self.longWait)
                selectWin = win32gui.GetForegroundWindow()
            win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

            img = ImageGrab.grabclipboard()
            img.save(self.imageLoc + filename + str(channel) + '1.png')
            img=openpyxl.drawing.image.Image(self.imageLoc + filename + str(channel) + '1.png')
            img.anchor = self.imageCol[channel * 2] + str(self.rowToImage)
            self.wsI.add_image(img)
            time.sleep(self.longWait)

        # Take second image
        pyautogui.moveTo(177, 536)
        pyautogui.click(button='right')
        time.sleep(self.mediumWait)
        self.click(259,551)
        time.sleep(self.mediumWait)
        pyautogui.keyDown('alt')
        time.sleep(self.shortWait)
        pyautogui.press('c')
        time.sleep(self.mediumWait)
        pyautogui.keyUp('alt')
        time.sleep(self.mediumWait)
        selectWin = win32gui.GetForegroundWindow()
        while win32gui.GetWindowText(selectWin) != "":
                time.sleep(self.longWait)
                selectWin = win32gui.GetForegroundWindow()
        win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

        img = ImageGrab.grabclipboard()
        img.save(self.imageLoc + filename + str(channel) + '2.png')
        img=openpyxl.drawing.image.Image(self.imageLoc + filename + str(channel) + '2.png')
        img.anchor = self.imageCol[1 + channel * 2] + str(self.rowToImage)
        self.wsI.add_image(img)
        time.sleep(self.longWait)

    def fitting(self, file, channel):
        # Fit
        self.click(137,32)
        time.sleep(self.mediumWait)
        selectWin = win32gui.GetForegroundWindow()
        win32gui.ShowWindow(selectWin, win32con.SW_MAXIMIZE)
        time.sleep(self.longWait)

        # Turn off G2(0)
        if (channel > 0) or (file > 0):
            self.click(117,245)
            time.sleep(self.shortWait)

        # Reset data
        self.click(36,185)
        self.click(36,185)
        time.sleep(self.mediumWait)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.shortWait)
        pyautogui.press('backspace')
        time.sleep(self.shortWait)
        pyautogui.press('1')
        self.click(36,206)
        self.click(36,206)
        time.sleep(self.mediumWait)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.shortWait)
        pyautogui.press('backspace')
        time.sleep(self.shortWait)
        pyautogui.press('1')
        time.sleep(self.shortWait)
        self.click(36,223)
        self.click(36,223)
        time.sleep(self.mediumWait)
        pyautogui.press('1')

        # Press fit
        self.click(244,168)

        # Close the pop up about error
        time.sleep(self.simulationWait)
        selectWin = win32gui.GetForegroundWindow()
        while win32gui.GetWindowText(selectWin) != "":
            time.sleep(self.longWait)
            selectWin = win32gui.GetForegroundWindow()
        win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

        self.click(125,600)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.mediumWait)
        pyautogui.hotkey('ctrl', 'c')
        data = pyperclip.paste()

        temp = re.search('Background(.+?)\r', data)
        self.ws['B' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('G1\(0\)(.+?)\r', data)
        self.ws['C' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('D1 \(in um2\/s\)(.+?)\r', data)
        self.ws['D' + str(self.rowToWrite)] = float(temp.group(1))
        if re.findall(r'\d+', data)[0] == '7':
            self.ws['B' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['C' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['D' + str(self.rowToWrite)].fill = self.fill_yellow
        elif re.findall(r'\d+', data)[0] == '8':
            self.ws['B' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['C' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['D' + str(self.rowToWrite)].fill = self.fill_red

        # Unlock G2(0)
        self.click(117,245)
        time.sleep(self.mediumWait)
        if (channel == 0) and (file == 0):
            self.click(72,245)
            self.click(72,264)

        # Set G2 values to 1
        self.click(36,244)
        self.click(36,244)
        time.sleep(self.mediumWait)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.mediumWait)
        pyautogui.press('1')
        self.click(36,263)
        self.click(36,263)
        time.sleep(self.mediumWait)
        pyautogui.press('1')
        time.sleep(self.mediumWait)

        # Press fit
        self.click(244,168)

        # Close the pop up about error
        time.sleep(self.simulationWait)
        selectWin = win32gui.GetForegroundWindow()
        while win32gui.GetWindowText(selectWin) != "":
            time.sleep(self.longWait)
            selectWin = win32gui.GetForegroundWindow()
        win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

        self.click(125,600)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(self.mediumWait)
        pyautogui.hotkey('ctrl', 'c')
        data = pyperclip.paste()

        temp = re.search('Background(.+?)\r', data)
        self.ws['F' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('G1\(0\)(.+?)\r', data)
        self.ws['G' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('D1 \(in um2\/s\)(.+?)\r', data)
        self.ws['H' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('Fraction vesicle(.+?)\r', data)
        self.ws['I' + str(self.rowToWrite)] = float(temp.group(1))
        temp = re.search('D2 \(in um2\/s\)(.+?)\r', data)
        self.ws['J' + str(self.rowToWrite)] = float(temp.group(1))
        if re.findall(r'\d+', data)[0] == '7':
            self.ws['F' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['G' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['H' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['I' + str(self.rowToWrite)].fill = self.fill_yellow
            self.ws['J' + str(self.rowToWrite)].fill = self.fill_yellow
        elif re.findall(r'\d+', data)[0] == '8':
            self.ws['F' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['G' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['H' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['I' + str(self.rowToWrite)].fill = self.fill_red
            self.ws['J' + str(self.rowToWrite)].fill = self.fill_red
        time.sleep(self.longWait)
        selectWin = win32gui.GetForegroundWindow()
        while win32gui.GetWindowText(selectWin) != "2D-ICS":
            time.sleep(self.longWait)
            selectWin = win32gui.GetForegroundWindow()
        win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

    def startSimFCS(self):
        self.fileLoc = self.fileLoc + '/'
        self.excelFileLoc = self.excelFileDir + '/' + self.excelFileName.text() + '.xlsx'
        self.imageLoc = self.imageLoc + '/'

        self.shortWait = float(self.shortWaitBox.text())
        self.mediumWait = float(self.mediumWaitBox.text())
        self.longWait = float(self.longWaitBox.text())
        self.simulationWait = float(self.simulationWaitBox.text())

        foldersInLoc = []

        files = os.listdir(self.fileLoc)
        for file in files:
            if file.endswith(".tif.frames"): foldersInLoc.append(file)

        wb = Workbook() #load_workbook(excelFileLoc)
        wb.create_sheet(title = "Ch 1")
        wb.create_sheet(title = "Ch 2")
        wb.create_sheet(title = "Ch1-Ch2 (Bcc map)")
        wb.create_sheet(title = "Ch2-Ch1 (B1-B2 map)")
        self.wsI = wb.create_sheet(title = "Images")
        sheetNames = ["Ch 1", "Ch 2", "Ch1-Ch2 (Bcc map)", "Ch2-Ch1 (B1-B2 map)"]
        for n in range(4):
            ws = ws = wb[sheetNames[n]]
            ws['A1'] = "Filename"
            ws['B1'] = "Background"
            ws['C1'] = "G1(0)"
            ws['D1'] = "D1 (in um2/s)"
            ws['F1'] = "Background"
            ws['G1'] = "G1(0)"
            ws['H1'] = "D1 (in um2/s)"
            ws['I1'] = "Fraction vesicle"
            ws['J1'] = "D2 (in um2/s)"

        self.imageCol = ['C', 'H', 'M', 'R', '', 'W', '', 'AB']

        subprocess.Popen(self.simFCSLoc)
        time.sleep(self.simulationWait)
        preScreen = win32gui.GetForegroundWindow()
        win32gui.ShowWindow(preScreen, win32con.SW_MAXIMIZE)
        pyautogui.press('enter')
        time.sleep(self.longWait)
        for n in range(len(foldersInLoc)):
            fileDir = self.fileLoc + foldersInLoc[n] + '/'

            tifFiles = []
            temp = os.listdir(fileDir)
            for file in temp:
                if file.endswith(".tif"): tifFiles.append(file)

            # Skip cases where there's not 200 tif files
            if len(tifFiles) != 200:
                continue
            
            temp = re.search(foldersInLoc[n][:-11] + '_(.+?)T100.tif', tifFiles[99])
            firstGroup = '*' + temp.group(1) + '*'
            temp = re.search(foldersInLoc[n][:-11] + '_(.+?)T100.tif', tifFiles[199])
            secondGroup = '*' + temp.group(1) + '*'

            mainWin = win32gui.GetForegroundWindow()
            win32gui.ShowWindow(mainWin, win32con.SW_MAXIMIZE)

            # Open window to upload second channel
            self.selectFiles()

            selectWin = win32gui.GetForegroundWindow()
            win32gui.ShowWindow(selectWin, win32con.SW_MAXIMIZE)
            time.sleep(self.mediumWait)

            # Select second channel
            self.selectGroup(fileDir, secondGroup)
            time.sleep(self.mediumWait)

            # Close the pop up
            selectWin = win32gui.GetForegroundWindow()
            while win32gui.GetWindowText(selectWin) != "File header":
                time.sleep(self.longWait)
                selectWin = win32gui.GetForegroundWindow()
            win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

            # Let it load
            time.sleep(self.longWait)

            # Exchange channels
            self.click(30,10)
            self.exchangeChannel()
            time.sleep(self.mediumWait)

            # Open window to upload another channel
            self.click(30,10)
            self.selectFiles()

            selectWin = win32gui.GetForegroundWindow()
            win32gui.ShowWindow(selectWin, win32con.SW_MAXIMIZE)
            time.sleep(self.mediumWait)

            # Select first channel
            self.selectGroup(fileDir, firstGroup)
            time.sleep(self.mediumWait)

            # Close the pop up
            time.sleep(self.mediumWait)
            selectWin = win32gui.GetForegroundWindow()
            while win32gui.GetWindowText(selectWin) != "File header":
                time.sleep(self.longWait)
                selectWin = win32gui.GetForegroundWindow()
            win32gui.PostMessage(selectWin,win32con.WM_CLOSE,0,0)

            # Let it load
            time.sleep(self.longWait)

            self.rowToWrite = 2 + n
            self.rowToImage = 2 + (n * 14)

            self.wsI['A' + str(self.rowToImage)] = foldersInLoc[n][:-11]
            # Loop the 4 channels
            for m in range(4):
                self.click(686,64)
                time.sleep(self.mediumWait)
                pyautogui.typewrite(sheetNames[m])
                self.ws = wb[sheetNames[m]]
                self.ws['A' + str(self.rowToWrite)] = foldersInLoc[n][:-11]
                # Tools -> RISC -> Substract moving average
                self.click(30,10)
                self.substractMovingAverage()

                self.takeImages(tifFiles[n][:-4], m)


                self.fitting(n, m)
                time.sleep(self.longWait)
                wb.save(self.excelFileLoc)

            wb.save(self.excelFileLoc)

if __name__ == '__main__':
    app = QApplication([])
    UI = SimFCSAutomator()
    UI.show()
    sys.exit(app.exec())
